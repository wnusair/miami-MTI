"""
API routes for data export and sensor data ingestion.
"""
from io import BytesIO
from datetime import datetime, timezone, timedelta
from flask import jsonify, request, send_file, abort
from flask_login import login_required, current_user
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from . import api_bp
from ...extensions import db
from ...models import SensorData


@api_bp.route('/sensor-data')
@login_required
def get_sensor_data():
    """Get sensor data for charts, with optional filtering."""
    # Get query parameters
    sensor_name = request.args.get('sensor_name')
    hours = request.args.get('hours', 1, type=int)
    limit = request.args.get('limit', 100, type=int)

    # Build query
    query = SensorData.query

    if sensor_name:
        query = query.filter(SensorData.sensor_name == sensor_name)

    # Filter by time range
    since = datetime.utcnow() - timedelta(hours=hours)
    query = query.filter(SensorData.timestamp >= since)

    # Order and limit
    data = query.order_by(SensorData.timestamp.desc()).limit(limit).all()

    return jsonify([item.to_dict() for item in reversed(data)])


@api_bp.route('/sensor-data/latest')
@login_required
def get_latest_sensor_data():
    """Get the latest reading for each sensor."""
    # Get distinct sensor names
    sensor_names = db.session.query(SensorData.sensor_name).distinct().all()
    sensor_names = [name[0] for name in sensor_names]

    latest_data = []
    for name in sensor_names:
        latest = SensorData.query.filter_by(sensor_name=name)\
            .order_by(SensorData.timestamp.desc()).first()
        if latest:
            latest_data.append(latest.to_dict())

    return jsonify(latest_data)


@api_bp.route('/sensor-data/stats')
@login_required
def get_sensor_stats():
    """Get statistics for dashboard KPIs."""
    hours = request.args.get('hours', 1, type=int)
    since = datetime.utcnow() - timedelta(hours=hours)

    # Get data from the last hour
    data = SensorData.query.filter(SensorData.timestamp >= since).all()

    if not data:
        return jsonify({
            'total_readings': 0,
            'sensor_count': 0,
            'avg_value': 0,
            'status_summary': {}
        })

    # Calculate stats
    sensor_names = set(d.sensor_name for d in data)
    status_counts = {}
    for d in data:
        status_counts[d.status] = status_counts.get(d.status, 0) + 1

    values = [d.value for d in data]
    avg_value = sum(values) / len(values) if values else 0

    return jsonify({
        'total_readings': len(data),
        'sensor_count': len(sensor_names),
        'avg_value': round(avg_value, 2),
        'status_summary': status_counts
    })


@api_bp.route('/export')
@login_required
def export_xlsx():
    """Export sensor data to XLSX file."""
    # Check permission
    if not current_user.can_export():
        abort(403, description="You do not have permission to export data.")
    
    # Get query parameters for date range
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = SensorData.query

    # Parse dates - handle both ISO format and datetime-local format
    if start_date:
        try:
            # Handle datetime-local format (2026-01-25T12:00)
            if 'T' in start_date and len(start_date) <= 16:
                start = datetime.strptime(start_date, '%Y-%m-%dT%H:%M')
            else:
                start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(SensorData.timestamp >= start)
        except (ValueError, AttributeError) as e:
            print(f"Start date parse error: {e}")

    if end_date:
        try:
            # Handle datetime-local format
            if 'T' in end_date and len(end_date) <= 16:
                end = datetime.strptime(end_date, '%Y-%m-%dT%H:%M')
            else:
                end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(SensorData.timestamp <= end)
        except (ValueError, AttributeError) as e:
            print(f"End date parse error: {e}")

    # Get data ordered by timestamp (newest first)
    data = query.order_by(SensorData.timestamp.desc()).all()

    # Miami University theme colors
    MIAMI_RED = 'C3142D'
    FOUNDATION_BLACK = '000000'
    CLINICAL_WHITE = 'FFFFFF'
    LIGHT_WASH = 'F5F7FA'
    BORDER_COLOR = 'E5E7EB'

    # Create workbook with openpyxl directly for better styling control
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sensor Data'

    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color=CLINICAL_WHITE)
    header_fill = PatternFill(start_color=MIAMI_RED, end_color=MIAMI_RED, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Arial', size=10, color=FOUNDATION_BLACK)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    # Alternating row colors
    row_fill_even = PatternFill(start_color=CLINICAL_WHITE, end_color=CLINICAL_WHITE, fill_type='solid')
    row_fill_odd = PatternFill(start_color=LIGHT_WASH, end_color=LIGHT_WASH, fill_type='solid')
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # Write headers
    headers = ['Timestamp', 'Sensor_ID', 'Value', 'Unit', 'Status']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, record in enumerate(data, 2):
        row_data = [
            record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            record.sensor_name,
            record.value,
            record.unit,
            record.status
        ]
        
        # Alternating row colors
        row_fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.fill = row_fill
            cell.border = thin_border
            
            # Special styling for Status column
            if col_idx == 5:  # Status column
                if value == 'OK':
                    cell.font = Font(name='Arial', size=10, color='059669', bold=True)
                elif value == 'WARNING':
                    cell.font = Font(name='Arial', size=10, color='D97706', bold=True)
                elif value == 'ERROR':
                    cell.font = Font(name='Arial', size=10, color=MIAMI_RED, bold=True)

    # Auto-adjust column widths
    column_widths = {'A': 22, 'B': 18, 'C': 12, 'D': 10, 'E': 12}
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Add a title row at the top
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value='MTI - Miami Telemetry Interface - Sensor Data Export')
    title_cell.font = Font(name='Arial', size=14, bold=True, color=MIAMI_RED)
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30
    
    # Add export info row
    ws.insert_rows(2)
    export_info = f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Records: {len(data)}'
    info_cell = ws.cell(row=2, column=1, value=export_info)
    info_cell.font = Font(name='Arial', size=9, italic=True, color='757575')
    ws.merge_cells('A2:E2')
    
    # Add empty row for spacing
    ws.insert_rows(3)
    ws.row_dimensions[3].height = 5

    # Write to BytesIO buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with timestamp
    filename = f'MTI_sensor_data_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@api_bp.route('/ingest', methods=['POST'])
def ingest_sensor_data():
    """Ingest sensor data from external sources."""
    data = request.get_json()

    if not data:
        return jsonify({'error': 'No data provided'}), 400

    # Handle single item or list
    items = data if isinstance(data, list) else [data]

    created = []
    for item in items:
        sensor_name = item.get('sensor_name')
        value = item.get('value')
        unit = item.get('unit', '')
        status = item.get('status', 'OK')

        if not sensor_name or value is None:
            continue

        try:
            value = float(value)
        except (TypeError, ValueError):
            continue

        sensor_data = SensorData(
            sensor_name=sensor_name,
            value=value,
            unit=unit,
            status=status
        )
        db.session.add(sensor_data)
        created.append(sensor_data.to_dict())

    db.session.commit()

    return jsonify({
        'created': len(created),
        'data': created
    }), 201
